from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from datetime import datetime
from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..database import get_database
from .auth import get_current_user
from ..models.user import UserRole

router = APIRouter(prefix="/api/export", tags=["export"])

def get_attendance_collection():
    db = get_database()
    return db["attendance"]

def get_users_collection():
    db = get_database()
    return db["users"]

def get_leaves_collection():
    db = get_database()
    return db["leaves"]

def get_overtime_collection():
    db = get_database()
    return db["overtime"]

# Styling
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

@router.get("/attendance")
async def export_attendance(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Export attendance report to Excel"""
    if current_user.get("role") not in [UserRole.SUPER_ADMIN.value, UserRole.HR_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Không có quyền xuất báo cáo")
    
    from calendar import monthrange
    
    attendance_col = get_attendance_collection()
    users_col = get_users_collection()
    
    # Get all active users
    users = await users_col.find({"status": "ACTIVE"}).to_list(1000)
    
    # Get attendance for the month
    first_day = datetime(year, month, 1)
    last_day = datetime(year, month, monthrange(year, month)[1], 23, 59, 59)
    
    attendance = await attendance_col.find({
        "timestamp": {"$gte": first_day, "$lte": last_day},
        "attendance_type": "CHECK_IN"
    }).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Chấm công T{month}/{year}"
    
    # Headers
    headers = ["STT", "Mã NV", "Họ tên", "Phòng ban", "Số ngày làm", "Đúng giờ", "Đi muộn", "Tỷ lệ %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER
    
    # Data rows
    row = 2
    for idx, user in enumerate(users, 1):
        user_id = str(user["_id"])
        user_attendance = [a for a in attendance if a["user_id"] == user_id]
        
        on_time = sum(1 for a in user_attendance if a["status"] == "ON_TIME")
        late = sum(1 for a in user_attendance if a["status"] == "LATE")
        total = len(user_attendance)
        rate = round(on_time / total * 100, 1) if total > 0 else 0
        
        data = [
            idx,
            user.get("employee_id", "N/A"),
            user.get("full_name", "N/A"),
            user.get("department", "N/A"),
            total,
            on_time,
            late,
            f"{rate}%"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"attendance_{year}_{month:02d}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/leaves")
async def export_leaves(
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Export leaves report to Excel"""
    if current_user.get("role") not in [UserRole.SUPER_ADMIN.value, UserRole.HR_MANAGER.value]:
        raise HTTPException(status_code=403, detail="Không có quyền xuất báo cáo")
    
    leaves_col = get_leaves_collection()
    
    # Get all leaves for the year
    leaves = await leaves_col.find({
        "start_date": {"$regex": f"^{year}"}
    }).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Nghỉ phép {year}"
    
    # Headers
    headers = ["STT", "Nhân viên", "Loại", "Từ ngày", "Đến ngày", "Số ngày", "Trạng thái", "Người duyệt"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER
    
    # Data rows
    for idx, leave in enumerate(leaves, 1):
        row = idx + 1
        data = [
            idx,
            leave.get("user_name", "N/A"),
            leave.get("leave_type", "N/A"),
            leave.get("start_date", "N/A"),
            leave.get("end_date", "N/A"),
            leave.get("days", 0),
            leave.get("status", "N/A"),
            leave.get("approved_by", "")
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    for i, width in enumerate([8, 25, 15, 12, 12, 10, 12, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=leaves_{year}.xlsx"}
    )

@router.get("/overtime")
async def export_overtime(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Export overtime report to Excel"""
    if current_user.get("role") not in [UserRole.SUPER_ADMIN.value, UserRole.HR_MANAGER.value, UserRole.ACCOUNTANT.value]:
        raise HTTPException(status_code=403, detail="Không có quyền xuất báo cáo")
    
    ot_col = get_overtime_collection()
    
    # Get all approved OT for the month
    date_pattern = f"{year}-{month:02d}"
    ots = await ot_col.find({
        "date": {"$regex": f"^{date_pattern}"},
        "status": "APPROVED"
    }).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"OT T{month}/{year}"
    
    # Headers
    headers = ["STT", "Nhân viên", "Phòng ban", "Ngày", "Từ", "Đến", "Số giờ", "Lý do"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER
    
    # Data rows
    for idx, ot in enumerate(ots, 1):
        row = idx + 1
        data = [
            idx,
            ot.get("user_name", "N/A"),
            ot.get("user_department", "N/A"),
            ot.get("date", "N/A"),
            ot.get("start_time", ""),
            ot.get("end_time", ""),
            ot.get("hours", 0),
            ot.get("reason", "")
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
    
    # Add summary
    total_row = len(ots) + 3
    ws.cell(row=total_row, column=6, value="Tổng cộng:").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(o.get("hours", 0) for o in ots)).font = Font(bold=True)
    
    for i, width in enumerate([8, 25, 20, 12, 8, 8, 10, 30], 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=overtime_{year}_{month:02d}.xlsx"}
    )
